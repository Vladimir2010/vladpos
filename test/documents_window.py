import sqlite3
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QPushButton,
    QComboBox, QLineEdit, QFrame, QFileDialog
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook, load_workbook

DB_PATH = "database/pos_system.db"


class DocumentsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Documents Window")
        self.setGeometry(100, 100, 800, 600)

        main_layout = QVBoxLayout()

        first_frame = QFrame()
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Products", "Total Amount", "Amount Paid", "Change", "Payment Type", "Cash Register Numbers"])
        first_layout.addWidget(self.table)

        self.load_documents()

        second_frame = QFrame()
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        buttons_layout = QHBoxLayout()

        self.edit_button = QPushButton("Edit Document")
        self.edit_button.clicked.connect(self.open_edit_document_window)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Delete Document")
        self.delete_button.clicked.connect(self.delete_document)
        buttons_layout.addWidget(self.delete_button)

        self.refresh_button = QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.load_documents)
        buttons_layout.addWidget(self.refresh_button)

        self.export_button = QPushButton("Export to XLSX")
        self.export_button.clicked.connect(self.export_to_xlsx)
        buttons_layout.addWidget(self.export_button)

        self.import_button = QPushButton("Import from XLSX")
        self.import_button.clicked.connect(self.import_from_xlsx)
        buttons_layout.addWidget(self.import_button)

        second_layout.addLayout(buttons_layout)

        # Add these lines in the __init__ method to create filter inputs
        self.filter_frame = QFrame()
        self.filter_layout = QVBoxLayout()
        self.filter_frame.setLayout(self.filter_layout)

        self.filter_products_input = QLineEdit()
        self.filter_products_input.setPlaceholderText("Filter by Products")
        self.filter_products_input.textChanged.connect(self.filter_documents)
        self.filter_layout.addWidget(self.filter_products_input)

        self.filter_id_input = QLineEdit()
        self.filter_id_input.setPlaceholderText("Filter by ID")
        self.filter_id_input.textChanged.connect(self.filter_documents)
        self.filter_layout.addWidget(self.filter_id_input)

        first_layout.addWidget(self.filter_frame)

        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)

        self.setLayout(main_layout)

    def load_documents(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute(
            'SELECT id, products, total_amount, amount_paid, change, payment_type, cash_register_numbers FROM documents')
        documents = cursor.fetchall()

        self.table.setRowCount(len(documents))
        for row_idx, document in enumerate(documents):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(document[0])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(document[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(document[2])))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(document[3])))
            self.table.setItem(row_idx, 4, QTableWidgetItem(str(document[4])))
            self.table.setItem(row_idx, 5, QTableWidgetItem(document[5]))
            self.table.setItem(row_idx, 6, QTableWidgetItem(document[6]))

        connection.close()

    def open_edit_document_window(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.edit_document_window = EditDocumentWindow(self.table, current_row)
            self.edit_document_window.exec()

    def delete_document(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            document_id = self.table.item(current_row, 0).text()
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM documents WHERE id = ?', (document_id,))
            connection.commit()
            connection.close()
            self.table.removeRow(current_row)
            self.load_documents()

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Documents"

            headers = ["ID", "Products", "Total Amount", "Amount Paid", "Change", "Payment Type",
                       "Cash Register Numbers"]
            sheet.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            workbook.save(file_path)

    def import_from_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            self.table.setRowCount(0)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for column, value in enumerate(row):
                    self.table.setItem(row_position, column, QTableWidgetItem(str(value)))

            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM documents')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cursor.execute('''
                    INSERT INTO documents (id, products, total_amount, amount_paid, change, payment_type, cash_register_numbers) VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', row)
            connection.commit()
            connection.close()

    # Add this method to filter the documents table
    def filter_documents(self):
        products_filter = self.filter_products_input.text().lower()
        id_filter = self.filter_id_input.text().lower()

        for row in range(self.table.rowCount()):
            products_item = self.table.item(row, 1)
            id_item = self.table.item(row, 0)
            products_match = products_filter in products_item.text().lower() if products_item else False
            id_match = id_filter in id_item.text().lower() if id_item else False
            self.table.setRowHidden(row, not (products_match and id_match))


class EditDocumentWindow(QDialog):
    def __init__(self, table, row):
        super().__init__()
        self.setWindowTitle("Edit Document")
        self.setGeometry(100, 100, 400, 400)
        self.table = table
        self.row = row

        layout = QVBoxLayout()

        self.id_input = QLineEdit()
        self.products_input = QLineEdit()
        self.total_amount_input = QLineEdit()
        self.amount_paid_input = QLineEdit()
        self.change_input = QLineEdit()
        self.payment_type_input = QLineEdit()
        self.cash_register_numbers_input = QLineEdit()

        layout.addWidget(QLabel("ID"))
        layout.addWidget(self.id_input)
        layout.addWidget(QLabel("Products"))
        layout.addWidget(self.products_input)
        layout.addWidget(QLabel("Total Amount"))
        layout.addWidget(self.total_amount_input)
        layout.addWidget(QLabel("Amount Paid"))
        layout.addWidget(self.amount_paid_input)
        layout.addWidget(QLabel("Change"))
        layout.addWidget(self.change_input)
        layout.addWidget(QLabel("Payment Type"))
        layout.addWidget(self.payment_type_input)
        layout.addWidget(QLabel("Cash Register Numbers"))
        layout.addWidget(self.cash_register_numbers_input)

        self.load_document()

        edit_button = QPushButton("Edit")
        edit_button.clicked.connect(self.edit_document)
        layout.addWidget(edit_button)

        self.setLayout(layout)

    def load_document(self):
        self.id_input.setText(self.table.item(self.row, 0).text())
        self.products_input.setText(self.table.item(self.row, 1).text())
        self.total_amount_input.setText(self.table.item(self.row, 2).text())
        self.amount_paid_input.setText(self.table.item(self.row, 3).text())
        self.change_input.setText(self.table.item(self.row, 4).text())
        self.payment_type_input.setText(self.table.item(self.row, 5).text())
        self.cash_register_numbers_input.setText(self.table.item(self.row, 6).text())

    def edit_document(self):
        document_id = self.id_input.text()
        products = self.products_input.text()
        total_amount = self.total_amount_input.text()
        amount_paid = self.amount_paid_input.text()
        change = self.change_input.text()
        payment_type = self.payment_type_input.text()
        cash_register_numbers = self.cash_register_numbers_input.text()

        self.table.setItem(self.row, 0, QTableWidgetItem(document_id))
        self.table.setItem(self.row, 1, QTableWidgetItem(products))
        self.table.setItem(self.row, 2, QTableWidgetItem(total_amount))
        self.table.setItem(self.row, 3, QTableWidgetItem(amount_paid))
        self.table.setItem(self.row, 4, QTableWidgetItem(change))
        self.table.setItem(self.row, 5, QTableWidgetItem(payment_type))
        self.table.setItem(self.row, 6, QTableWidgetItem(cash_register_numbers))

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            UPDATE documents SET products = ?, total_amount = ?, amount_paid = ?, change = ?, payment_type = ?, cash_register_numbers = ? WHERE id = ?
        ''', (products, total_amount, amount_paid, change, payment_type, cash_register_numbers, document_id))
        connection.commit()
        connection.close()

        self.accept()