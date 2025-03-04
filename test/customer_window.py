import sqlite3
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QPushButton,
    QComboBox, QLineEdit, QFrame, QFileDialog
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook, load_workbook

DB_DIR = "database"
DB_PATH = "database/pos_system.db"

class CustomersWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Контрагенти")
        self.setGeometry(100, 100, 600, 400)

        main_layout = QVBoxLayout()

        first_frame = QFrame()
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Име на фирма", "Булстат", "ДДС Номер", "Адрес", "МОЛ", "Тел. Номер"])
        first_layout.addWidget(self.table)

        self.load_customers()

        second_frame = QFrame()
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        buttons_layout = QHBoxLayout()
        self.add_button = QPushButton("Нов")
        self.add_button.clicked.connect(self.open_add_customer_window)
        buttons_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Редактирай")
        self.edit_button.clicked.connect(self.open_edit_customer_window)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Изтрий")
        self.delete_button.clicked.connect(self.delete_customer)
        buttons_layout.addWidget(self.delete_button)

        self.refresh_button = QPushButton("Обнови")
        self.refresh_button.clicked.connect(self.load_customers)
        buttons_layout.addWidget(self.refresh_button)

        self.export_button = QPushButton("Експортиране в XLSX")
        self.export_button.clicked.connect(self.export_to_xlsx)
        buttons_layout.addWidget(self.export_button)

        self.import_button = QPushButton("Импортиране от XLSX")
        self.import_button.clicked.connect(self.import_from_xlsx)
        buttons_layout.addWidget(self.import_button)

        second_layout.addLayout(buttons_layout)

        # Add these lines in the __init__ method to create filter inputs
        self.filter_frame = QFrame()
        self.filter_layout = QVBoxLayout()
        self.filter_frame.setLayout(self.filter_layout)

        self.filter_name_input = QLineEdit()
        self.filter_name_input.setPlaceholderText("Филтър по Име на фирма")
        self.filter_name_input.textChanged.connect(self.filter_customers)
        self.filter_layout.addWidget(self.filter_name_input)

        self.filter_eik_input = QLineEdit()
        self.filter_eik_input.setPlaceholderText("Филтър по Булстат")
        self.filter_eik_input.textChanged.connect(self.filter_customers)
        self.filter_layout.addWidget(self.filter_eik_input)

        self.filter_phone_input = QLineEdit()
        self.filter_phone_input.setPlaceholderText("Филтър по Тел. Номер")
        self.filter_phone_input.textChanged.connect(self.filter_customers)
        self.filter_layout.addWidget(self.filter_phone_input)

        first_layout.addWidget(self.filter_frame)

        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)

        self.setLayout(main_layout)

    def load_customers(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT name, eik, dds, address, mol, phone FROM customers')
        customers = cursor.fetchall()

        self.table.setRowCount(len(customers))
        for row_idx, customer in enumerate(customers):
            self.table.setItem(row_idx, 0, QTableWidgetItem(customer[0]))
            self.table.setItem(row_idx, 1, QTableWidgetItem(customer[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem(customer[2]))
            self.table.setItem(row_idx, 3, QTableWidgetItem(customer[3]))
            self.table.setItem(row_idx, 4, QTableWidgetItem(customer[4]))
            self.table.setItem(row_idx, 5, QTableWidgetItem(customer[5]))

        connection.close()

    def open_add_customer_window(self):
        self.add_customer_window = AddCustomerWindow(self.table)
        self.add_customer_window.exec()

    def open_edit_customer_window(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.edit_customer_window = EditCustomerWindow(self.table, current_row)
            self.edit_customer_window.exec()

    def delete_customer(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            eik = self.table.item(current_row, 1).text()
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM customers WHERE eik = ?', (eik,))
            connection.commit()
            connection.close()
            self.table.removeRow(current_row)
            self.load_customers()

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Контрагенти"

            headers = ["Име нафирма", "Булстат", "ДДС Номер", "Адрес", "МОЛ", "Тел. номер"]
            sheet.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            workbook.save(file_path)

    def import_from_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            self.table.setRowCount(0)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for column, value in enumerate(row):
                    self.table.setItem(row_position, column, QTableWidgetItem(str(value)))

            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM customers')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cursor.execute('''
                    INSERT INTO customers (name, eik, dds, address, mol, phone) VALUES (?, ?, ?, ?, ?, ?)
                ''', row)
            connection.commit()
            connection.close()

    # Add this method to filter the customers table
    def filter_customers(self):
        name_filter = self.filter_name_input.text().lower()
        eik_filter = self.filter_eik_input.text().lower()
        phone_filter = self.filter_phone_input.text().lower()

        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            eik_item = self.table.item(row, 1)
            phone_item = self.table.item(row, 5)
            name_match = name_filter in name_item.text().lower() if name_item else False
            eik_match = eik_filter in eik_item.text().lower() if eik_item else False
            phone_match = phone_filter in phone_item.text().lower() if phone_item else False
            self.table.setRowHidden(row, not (name_match and eik_match and phone_match))

class AddCustomerWindow(QDialog):
    def __init__(self, table):
        super().__init__()
        self.setWindowTitle("Нов контрагент")
        self.setGeometry(100, 100, 300, 300)
        self.table = table

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.eik_input = QLineEdit()
        self.dds_input = QLineEdit()
        self.address_input = QLineEdit()
        self.mol_input = QLineEdit()
        self.phone_input = QLineEdit()

        layout.addWidget(QLabel("Име на фирма"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Булстат"))
        layout.addWidget(self.eik_input)
        layout.addWidget(QLabel("ДДС номер"))
        layout.addWidget(self.dds_input)
        layout.addWidget(QLabel("Адрес"))
        layout.addWidget(self.address_input)
        layout.addWidget(QLabel("МОЛ"))
        layout.addWidget(self.mol_input)
        layout.addWidget(QLabel("Тел. номер"))
        layout.addWidget(self.phone_input)

        add_button = QPushButton("Добави")
        add_button.clicked.connect(self.add_customer)
        layout.addWidget(add_button)

        self.setLayout(layout)

    def add_customer(self):
        name = self.name_input.text()
        eik = self.eik_input.text()
        dds = self.dds_input.text()
        address = self.address_input.text()
        mol = self.mol_input.text()
        phone = self.phone_input.text()

        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        self.table.setItem(row_position, 0, QTableWidgetItem(name))
        self.table.setItem(row_position, 1, QTableWidgetItem(eik))
        self.table.setItem(row_position, 2, QTableWidgetItem(dds))
        self.table.setItem(row_position, 3, QTableWidgetItem(address))
        self.table.setItem(row_position, 4, QTableWidgetItem(mol))
        self.table.setItem(row_position, 5, QTableWidgetItem(phone))

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            INSERT INTO customers (name, eik, dds, address, mol, phone) VALUES (?, ?, ?, ?, ?, ?)
        ''', (name, eik, dds, address, mol, phone))
        connection.commit()
        connection.close()

        self.accept()

class EditCustomerWindow(QDialog):
    def __init__(self, table, row):
        super().__init__()
        self.setWindowTitle("Редактиране на контрагент")
        self.setGeometry(100, 100, 300, 300)
        self.table = table
        self.row = row

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.eik_input = QLineEdit()
        self.dds_input = QLineEdit()
        self.address_input = QLineEdit()
        self.mol_input = QLineEdit()
        self.phone_input = QLineEdit()

        layout.addWidget(QLabel("Име на фирма"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Булстат"))
        layout.addWidget(self.eik_input)
        layout.addWidget(QLabel("ДДС номер"))
        layout.addWidget(self.dds_input)
        layout.addWidget(QLabel("Адрес"))
        layout.addWidget(self.address_input)
        layout.addWidget(QLabel("МОЛ"))
        layout.addWidget(self.mol_input)
        layout.addWidget(QLabel("Тел. номер"))
        layout.addWidget(self.phone_input)

        self.load_customer()

        edit_button = QPushButton("Запазване")
        edit_button.clicked.connect(self.edit_customer)
        layout.addWidget(edit_button)

        self.setLayout(layout)

    def load_customer(self):
        self.name_input.setText(self.table.item(self.row, 0).text())
        self.eik_input.setText(self.table.item(self.row, 1).text())
        self.dds_input.setText(self.table.item(self.row, 2).text())
        self.address_input.setText(self.table.item(self.row, 3).text())
        self.mol_input.setText(self.table.item(self.row, 4).text())
        self.phone_input.setText(self.table.item(self.row, 5).text())

    def edit_customer(self):
        name = self.name_input.text()
        eik = self.eik_input.text()
        dds = self.dds_input.text()
        address = self.address_input.text()
        mol = self.mol_input.text()
        phone = self.phone_input.text()

        self.table.setItem(self.row, 0, QTableWidgetItem(name))
        self.table.setItem(self.row, 1, QTableWidgetItem(eik))
        self.table.setItem(self.row, 2, QTableWidgetItem(dds))
        self.table.setItem(self.row, 3, QTableWidgetItem(address))
        self.table.setItem(self.row, 4, QTableWidgetItem(mol))
        self.table.setItem(self.row, 5, QTableWidgetItem(phone))

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            UPDATE customers SET name = ?, eik = ?, dds = ?, address = ?, mol = ?, phone = ? WHERE eik = ?
        ''', (name, eik, dds, address, mol, phone, self.table.item(self.row, 1).text()))
        connection.commit()
        connection.close()

        self.accept()