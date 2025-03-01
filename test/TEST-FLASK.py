# import sys
# from PyQt6.QtWidgets import (
#     QApplication, QMainWindow, QFrame, QVBoxLayout, QHBoxLayout, QWidget, QMenu,
#     QSizePolicy, QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
#     QHeaderView, QDialog, QDialogButtonBox, QCalendarWidget, QTextEdit
# )
# from PyQt6.QtGui import QIcon, QAction, QPixmap
# from PyQt6.QtCore import Qt
#
# class CashRegisterWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Cash Register")
#         self.setGeometry(100, 100, 600, 400)
#         main_layout = QVBoxLayout()
#
#         # First Frame
#         first_frame = QFrame()
#         first_frame.setFrameShape(QFrame.Shape.NoFrame)
#         first_layout = QVBoxLayout()
#         first_frame.setLayout(first_layout)
#
#         first_title = QLabel("First Title")
#         first_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         first_layout.addWidget(first_title)
#
#         first_row_layout = QHBoxLayout()
#         first_label = QLabel("Label:")
#         first_input = QLineEdit()
#         plus_button = QPushButton("+")
#         minus_button = QPushButton("-")
#         first_row_layout.addWidget(first_label)
#         first_row_layout.addWidget(first_input)
#         first_row_layout.addWidget(plus_button)
#         first_row_layout.addWidget(minus_button)
#         first_layout.addLayout(first_row_layout)
#
#         # Second Frame
#         second_frame = QFrame()
#         second_frame.setFrameShape(QFrame.Shape.NoFrame)
#         second_layout = QVBoxLayout()
#         second_frame.setLayout(second_layout)
#
#         second_title = QLabel("Second Title")
#         second_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         second_layout.addWidget(second_title)
#
#         second_row_layout1 = QHBoxLayout()
#         second_label1 = QLabel("Label 1:")
#         second_label2 = QLabel("Label 2:")
#         second_row_layout1.addWidget(second_label1)
#         second_row_layout1.addWidget(second_label2)
#         second_layout.addLayout(second_row_layout1)
#
#         second_row_layout2 = QHBoxLayout()
#         second_input1 = QCalendarWidget()
#         second_input2 = QCalendarWidget()
#         second_row_layout2.addWidget(second_input1)
#         second_row_layout2.addWidget(second_input2)
#         second_layout.addLayout(second_row_layout2)
#
#         second_button = QPushButton("Button")
#         second_layout.addWidget(second_button)
#
#         # Third Frame
#         third_frame = QFrame()
#         third_frame.setFrameShape(QFrame.Shape.NoFrame)
#         third_layout = QVBoxLayout()
#         third_frame.setLayout(third_layout)
#
#         third_title = QLabel("Third Title")
#         third_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         third_layout.addWidget(third_title)
#
#         third_row_layout1 = QHBoxLayout()
#         third_label1 = QLabel("Label 1:")
#         third_label2 = QLabel("Label 2:")
#         third_row_layout1.addWidget(third_label1)
#         third_row_layout1.addWidget(third_label2)
#         third_layout.addLayout(third_row_layout1)
#
#         third_row_layout2 = QHBoxLayout()
#         third_input1 = QCalendarWidget()
#         third_input2 = QCalendarWidget()
#         third_row_layout2.addWidget(third_input1)
#         third_row_layout2.addWidget(third_input2)
#         third_layout.addLayout(third_row_layout2)
#
#         third_button = QPushButton("Button")
#         third_layout.addWidget(third_button)
#
#         # Fourth Frame
#         fourth_frame = QFrame()
#         fourth_frame.setFrameShape(QFrame.Shape.NoFrame)
#         fourth_layout = QVBoxLayout()
#         fourth_frame.setLayout(fourth_layout)
#
#         fourth_title = QLabel("Fourth Title")
#         fourth_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         fourth_layout.addWidget(fourth_title)
#
#         fourth_button = QPushButton("Button")
#         fourth_layout.addWidget(fourth_button)
#
#         # Add frames to main layout
#         main_layout.addWidget(first_frame)
#         main_layout.addWidget(second_frame)
#         main_layout.addWidget(third_frame)
#         main_layout.addWidget(fourth_frame)
#
#         self.setLayout(main_layout)
#
# class SpravkiDocumentsWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Spravki Documents")
#         self.setGeometry(100, 100, 400, 300)
#         layout = QVBoxLayout()
#         label = QLabel("Spravki Documents Window")
#         layout.addWidget(label)
#         button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
#         button_box.accepted.connect(self.accept)
#         layout.addWidget(button_box)
#         self.setLayout(layout)
#
# class SpravkiClientsWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Spravki Clients")
#         self.setGeometry(100, 100, 400, 300)
#         layout = QVBoxLayout()
#         label = QLabel("Spravki Clients Window")
#         layout.addWidget(label)
#         button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
#         button_box.accepted.connect(self.accept)
#         layout.addWidget(button_box)
#         self.setLayout(layout)
#
# class SpravkiSalesWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Spravki Sales")
#         self.setGeometry(100, 100, 400, 300)
#         layout = QVBoxLayout()
#         label = QLabel("Spravki Sales Window")
#         layout.addWidget(label)
#         button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
#         button_box.accepted.connect(self.accept)
#         layout.addWidget(button_box)
#         self.setLayout(layout)
#
# class SettingsWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Settings")
#         self.setGeometry(100, 100, 400, 300)
#         layout = QVBoxLayout()
#         label = QLabel("Settings Window")
#         layout.addWidget(label)
#         button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
#         button_box.accepted.connect(self.accept)
#         layout.addWidget(button_box)
#         self.setLayout(layout)
#
# class HelpWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Help")
#         self.setGeometry(100, 100, 400, 300)
#         layout = QVBoxLayout()
#         help_text = QTextEdit()
#         help_text.setReadOnly(True)
#         help_text.setText("This is the help text for the program. Here you can provide information about how to use the application, tips, and any other relevant information.")
#         layout.addWidget(help_text)
#         button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
#         button_box.accepted.connect(self.accept)
#         layout.addWidget(button_box)
#         self.setLayout(layout)
#
# class MainWindow(QMainWindow):
#     def __init__(self):
#         super().__init__()
#
#         self.setWindowTitle("Main Window with Three Horizontal Frames and Menubar")
#         self.setWindowIcon(QIcon('vladpos_logo.png'))
#         self.setGeometry(100, 100, 800, 600)
#         self.showMaximized()
#
#         # Set up the central widget and layout
#         central_widget = QWidget()
#         self.setCentralWidget(central_widget)
#         main_layout = QVBoxLayout()
#         central_widget.setLayout(main_layout)
#
#         # Create frames
#         frame1 = QFrame()
#         frame1.setFrameShape(QFrame.Shape.NoFrame)
#         frame1.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
#
#         frame2 = QFrame()
#         frame2.setFrameShape(QFrame.Shape.NoFrame)
#         frame2.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
#
#         frame3 = QFrame()
#         frame3.setFrameShape(QFrame.Shape.NoFrame)
#         frame3.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
#
#         # Add frames to layout
#         main_layout.addWidget(frame1, 1)
#         main_layout.addWidget(frame2, 5)
#         main_layout.addWidget(frame3, 2)
#
#         # Create menubar
#         menubar = self.menuBar()
#
#         # Create Cash Register menu
#         cash_register_menu = QMenu("Cash Register", self)
#         cash_register_menu.setIcon(QIcon('receipt.png'))
#         cash_register_action = QAction(QIcon('receipt.png'), 'Cash Register Action', self)
#         cash_register_action.triggered.connect(self.open_cash_register_window)
#         cash_register_menu.addAction(cash_register_action)
#         menubar.addMenu(cash_register_menu)
#
#         # Create Spravki menu
#         spravki_menu = QMenu("Spravki", self)
#         spravki_menu.setIcon(QIcon('paper.png'))
#
#         # Spravki Documents submenu
#         spravki_documents_action = QAction('Spravki Documents', self)
#         spravki_documents_action.triggered.connect(self.open_spravki_documents_window)
#         spravki_menu.addAction(spravki_documents_action)
#
#         # Spravki Clients submenu
#         spravki_clients_action = QAction('Spravki Clients', self)
#         spravki_clients_action.triggered.connect(self.open_spravki_clients_window)
#         spravki_menu.addAction(spravki_clients_action)
#
#         # Spravki Sales submenu
#         spravki_sales_action = QAction('Spravki Sales', self)
#         spravki_sales_action.triggered.connect(self.open_spravki_sales_window)
#         spravki_menu.addAction(spravki_sales_action)
#
#         menubar.addMenu(spravki_menu)
#
#         # Create Settings menu
#         settings_menu = QMenu("Settings", self)
#         settings_menu.setIcon(QIcon('setting.png'))
#         settings_action = QAction(QIcon('setting.png'), 'Settings Action', self)
#         settings_action.triggered.connect(self.open_settings_window)
#         settings_menu.addAction(settings_action)
#         menubar.addMenu(settings_menu)
#
#         # Create Help menu
#         help_menu = QMenu("Help", self)
#         help_menu.setIcon(QIcon('help.png'))
#         help_action = QAction(QIcon('help.png'), 'Help', self)
#         help_action.triggered.connect(self.open_help_window)
#         help_menu.addAction(help_action)
#         menubar.addMenu(help_menu)
#
#         # Frame 1 Contents
#         frame1_layout = QHBoxLayout()
#         frame1.setLayout(frame1_layout)
#
#         self.input1 = QLineEdit()
#         self.input1.setPlaceholderText("Input 1")
#         frame1_layout.addWidget(self.input1, 3)  # Biggest
#
#         self.input2 = QLineEdit()
#         self.input2.setPlaceholderText("Input 2")
#         frame1_layout.addWidget(self.input2, 1)  # Small
#
#         self.input3 = QLineEdit()
#         self.input3.setPlaceholderText("Input 3")
#         frame1_layout.addWidget(self.input3, 1)  # Like the second
#
#         self.popup_menu1 = QComboBox()
#         self.popup_menu1.addItems(["Option 1", "Option 2", "Option 3", "Option 4"])
#         frame1_layout.addWidget(self.popup_menu1, 2)  # A little bit more bigger
#
#         self.add_button = QPushButton("Add")
#         frame1_layout.addWidget(self.add_button, 2)  # Bigger
#
#         # Frame 2 Contents
#         self.table = QTableWidget()
#         self.table.setColumnCount(6)
#         self.table.setHorizontalHeaderLabels(["Input 1", "Input 2", "Input 3", "Popup Menu", "Multiplication", "Actions"])
#         self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
#         frame2_layout = QVBoxLayout()
#         frame2.setLayout(frame2_layout)
#         frame2_layout.addWidget(self.table)
#
#         # Frame 3 Contents
#         frame3_layout = QVBoxLayout()
#         frame3.setLayout(frame3_layout)
#
#         sub_frame1 = QFrame()
#         sub_frame1.setFrameShape(QFrame.Shape.NoFrame)
#         sub_frame1_layout = QVBoxLayout()
#         sub_frame1.setLayout(sub_frame1_layout)
#         image_label = QLabel()
#         image_label.setPixmap(QPixmap('path/to/image.png'))
#         image_label.setScaledContents(True)  # Scale the image to fit the label
#         image_label.setMaximumSize(200, 200)  # Set maximum size for the image label
#         sub_frame1_layout.addWidget(image_label, alignment=Qt.AlignmentFlag.AlignCenter)
#         image_text_label = QLabel("Image Label")
#         image_text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Center the text under the image
#         sub_frame1_layout.addWidget(image_text_label)
#
#         sub_frame2 = QFrame()
#         sub_frame2.setFrameShape(QFrame.Shape.NoFrame)
#         sub_frame2_layout = QVBoxLayout()
#         sub_frame2.setLayout(sub_frame2_layout)
#         popup_menu2 = QComboBox()
#         popup_menu2.addItems(["Option 1", "Option 2", "Option 3", "Option 4"])
#         sub_frame2_layout.addWidget(popup_menu2)
#         button1 = QPushButton("Button 1")
#         sub_frame2_layout.addWidget(button1)
#         button2 = QPushButton("Button 2")
#         sub_frame2_layout.addWidget(button2)
#
#         sub_frame3 = QFrame()
#         sub_frame3.setFrameShape(QFrame.Shape.NoFrame)
#         sub_frame3_layout = QVBoxLayout()
#         sub_frame3.setLayout(sub_frame3_layout)
#         input4 = QLineEdit()
#         input4.setPlaceholderText("Input 4")
#         sub_frame3_layout.addWidget(input4)
#         button3 = QPushButton("Button 3")
#         sub_frame3_layout.addWidget(button3)
#         button4 = QPushButton("Button 4")
#         sub_frame3_layout.addWidget(button4)
#
#         sub_frame4 = QFrame()
#         sub_frame4.setFrameShape(QFrame.Shape.NoFrame)
#         sub_frame4_layout = QVBoxLayout()
#         sub_frame4.setLayout(sub_frame4_layout)
#         input5 = QLineEdit()
#         input5.setPlaceholderText("Input 5")
#         sub_frame4_layout.addWidget(input5)
#         button5 = QPushButton("Button 5")
#         sub_frame4_layout.addWidget(button5)
#         button6 = QPushButton("Button 6")
#         sub_frame4_layout.addWidget(button6)
#
#         frame3_sub_layout = QHBoxLayout()
#         frame3_layout.addLayout(frame3_sub_layout)
#         frame3_sub_layout.addWidget(sub_frame1)
#         frame3_sub_layout.addWidget(sub_frame2)
#         frame3_sub_layout.addWidget(sub_frame3)
#         frame3_sub_layout.addWidget(sub_frame4)
#
#         # Set equal stretch for sub-frames
#         frame3_sub_layout.setStretch(0, 1)
#         frame3_sub_layout.setStretch(1, 1)
#         frame3_sub_layout.setStretch(2, 1)
#         frame3_sub_layout.setStretch(3, 1)
#
#         # Connect the add button to the function
#         self.add_button.clicked.connect(self.add_to_table)
#
#     def add_to_table(self):
#         row_position = self.table.rowCount()
#         self.table.insertRow(row_position)
#
#         input1_text = self.input1.text()
#         input2_text = self.input2.text()
#         input3_text = self.input3.text()
#         popup_text = self.popup_menu1.currentText()
#
#         # Add the input data to the table
#         self.table.setItem(row_position, 0, QTableWidgetItem(input1_text))
#         self.table.setItem(row_position, 1, QTableWidgetItem(input2_text))
#         self.table.setItem(row_position, 2, QTableWidgetItem(input3_text))
#         self.table.setItem(row_position, 3, QTableWidgetItem(popup_text))
#
#         # Calculate the multiplication of input2 and input3
#         try:
#             multiplication_result = float(input2_text) * float(input3_text)
#         except ValueError:
#             multiplication_result = "Error"
#         self.table.setItem(row_position, 4, QTableWidgetItem(str(multiplication_result)))
#
#         # Add a delete button with trash icon
#         delete_button = QPushButton()
#         delete_button.setIcon(QIcon('trash_icon.png'))
#         delete_button.clicked.connect(lambda: self.remove_row(row_position))
#         self.table.setCellWidget(row_position, 5, delete_button)
#
#         # Clear the inputs after adding to the table
#         self.input1.clear()
#         self.input2.clear()
#         self.input3.clear()
#         self.popup_menu1.setCurrentIndex(0)
#
#     def remove_row(self, row):
#         self.table.removeRow(row)
#
#     def open_cash_register_window(self):
#         self.cash_register_window = CashRegisterWindow()
#         self.cash_register_window.exec()
#
#     def open_spravki_documents_window(self):
#         self.spravki_documents_window = SpravkiDocumentsWindow()
#         self.spravki_documents_window.exec()
#
#     def open_spravki_clients_window(self):
#         self.spravki_clients_window = SpravkiClientsWindow()
#         self.spravki_clients_window.exec()
#
#     def open_spravki_sales_window(self):
#         self.spravki_sales_window = SpravkiSalesWindow()
#         self.spravki_sales_window.exec()
#
#     def open_settings_window(self):
#         self.settings_window = SettingsWindow()
#         self.settings_window.exec()
#
#     def open_help_window(self):
#         self.help_window = HelpWindow()
#         self.help_window.exec()
#
# if __name__ == "__main__":
#     app = QApplication(sys.argv)
#     window = MainWindow()
#     window.show()
#     sys.exit(app.exec())
import sys
import os
import sqlite3
import json
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon, QAction, QPixmap
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFrame, QVBoxLayout, QHBoxLayout, QWidget, QMenu,
    QSizePolicy, QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialog, QDialogButtonBox, QCalendarWidget, QTextEdit, QListWidget,
    QFileDialog, QAbstractItemView, QMessageBox, QTabWidget
)
from customer_window import CustomersWindow
from documents_window import DocumentsWindow
from invoices_window import InvoicesWindow

DB_DIR = "database"
# Функция за създаване на база данни и таблици
DB_PATH = "database/pos_system.db"

def create_database():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)
    # Проверка дали файлът за базата данни съществува
    if not os.path.exists(DB_PATH):
        open(DB_PATH, 'w').close()  # Създаване на празен файл за базата
        print("Създаден е нов файл за базата данни.")
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()

    # Create products table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            barcode TEXT NOT NULL,
            unit_price REAL NOT NULL
        )
    ''')

    # Create customers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            eik TEXT NOT NULL,
            dds TEXT NOT NULL,
            address TEXT NOT NULL,
            mol TEXT NOT NULL,
            phone TEXT NOT NULL
        )
    ''')

    # Create firm table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS firm (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            eik TEXT NOT NULL,
            dds TEXT NOT NULL,
            address TEXT NOT NULL,
            mol TEXT NOT NULL,
            phone TEXT NOT NULL
        )
    ''')

    # Create documents table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY,
            products TEXT NOT NULL,  -- JSON format
            total_amount REAL NOT NULL,
            amount_paid REAL NOT NULL,
            change REAL NOT NULL,
            payment_type TEXT NOT NULL,
            cash_register_numbers TEXT
        )
    ''')

    # Create invoices table with foreign keys referencing customers and firm
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            number INTEGER PRIMARY KEY,
            customer_ref_no INTEGER NOT NULL,
            selling_company_ref_no INTEGER NOT NULL,
            date TEXT NOT NULL,
            products TEXT NOT NULL,  -- JSON format
            total_amount REAL NOT NULL,
            amount_paid REAL NOT NULL,
            change REAL NOT NULL,
            payment_type TEXT NOT NULL,
            FOREIGN KEY (customer_ref_no) REFERENCES customers(id),
            FOREIGN KEY (selling_company_ref_no) REFERENCES firm(id)
        )
    ''')

    connection.commit()
    connection.close()

create_database()

class ProductsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Products Window")
        self.setGeometry(100, 100, 600, 400)

        main_layout = QVBoxLayout()

        # First Frame
        first_frame = QFrame()
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Name", "Barcode", "Price"])
        first_layout.addWidget(self.table)

        # Load products into the table
        self.load_products()

        # Second Frame
        second_frame = QFrame()
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        buttons_layout = QHBoxLayout()
        self.add_button = QPushButton("Add New Product")
        self.add_button.clicked.connect(self.open_add_product_window)
        buttons_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Edit Product")
        self.edit_button.clicked.connect(self.open_edit_product_window)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Delete Product")
        self.delete_button.clicked.connect(self.delete_product)
        buttons_layout.addWidget(self.delete_button)

        self.refresh_button = QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.load_products)
        buttons_layout.addWidget(self.refresh_button)

        self.export_button = QPushButton("Export to XLSX")
        self.export_button.clicked.connect(self.export_to_xlsx)
        buttons_layout.addWidget(self.export_button)

        self.import_button = QPushButton("Import from XLSX")
        self.import_button.clicked.connect(self.import_from_xlsx)
        buttons_layout.addWidget(self.import_button)

        second_layout.addLayout(buttons_layout)

        combo_layout = QHBoxLayout()
        self.combo1 = QComboBox()
        self.combo2 = QComboBox()
        combo_layout.addWidget(self.combo1)
        combo_layout.addWidget(self.combo2)
        second_layout.addLayout(combo_layout)

        # Third Frame

        # Add these lines in the __init__ method to create filter inputs
        self.filter_frame = QFrame()
        self.filter_layout = QHBoxLayout()
        self.filter_frame.setLayout(self.filter_layout)
        self.filter_name_input = QLineEdit()
        self.filter_name_input.setPlaceholderText("Filter by Name")
        self.filter_name_input.textChanged.connect(self.filter_products)
        self.filter_layout.addWidget(self.filter_name_input)
        self.filter_barcode_input = QLineEdit()
        self.filter_barcode_input.setPlaceholderText("Filter by Barcode")
        self.filter_barcode_input.textChanged.connect(self.filter_products)
        self.filter_layout.addWidget(self.filter_barcode_input)
        first_layout.addWidget(self.filter_frame)


        # Add frames to main layout
        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)
        main_layout.addWidget(self.filter_frame)



        self.setLayout(main_layout)

    def load_products(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT name, barcode, unit_price FROM products')
        products = cursor.fetchall()

        self.table.setRowCount(len(products))
        for row_idx, product in enumerate(products):
            self.table.setItem(row_idx, 0, QTableWidgetItem(product[0]))
            self.table.setItem(row_idx, 1, QTableWidgetItem(product[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(product[2])))

        connection.close()

    def open_add_product_window(self):
        self.add_product_window = AddProductWindow(self.table)
        self.add_product_window.exec()

    def open_edit_product_window(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.edit_product_window = EditProductWindow(self.table, current_row)
            self.edit_product_window.exec()

    def delete_product(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            barcode = self.table.item(current_row, 1).text()
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM products WHERE barcode = ?', (barcode,))
            connection.commit()
            connection.close()
            self.table.removeRow(current_row)
            self.load_products()

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Products"

            headers = ["Name", "Barcode", "Price"]
            sheet.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            workbook.save(file_path)

    def import_from_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            self.table.setRowCount(0)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for column, value in enumerate(row):
                    self.table.setItem(row_position, column, QTableWidgetItem(str(value)))

            # Optionally, update the database with the imported data
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM products')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cursor.execute('''
                    INSERT INTO products (name, barcode, unit_price) VALUES (?, ?, ?)
                ''', row)
            connection.commit()
            connection.close()

    # Add this method to filter the products table
    def filter_products(self):
        name_filter = self.filter_name_input.text().lower()
        barcode_filter = self.filter_barcode_input.text().lower()

        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            barcode_item = self.table.item(row, 1)
            name_match = name_filter in name_item.text().lower() if name_item else False
            barcode_match = barcode_filter in barcode_item.text().lower() if barcode_item else False
            self.table.setRowHidden(row, not (name_match and barcode_match))

class AddProductWindow(QDialog):
    def __init__(self, table):
        super().__init__()
        self.setWindowTitle("Add Product")
        self.setGeometry(100, 100, 300, 200)
        self.table = table

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.barcode_input = QLineEdit()
        self.price_input = QLineEdit()

        layout.addWidget(QLabel("Name"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Barcode"))
        layout.addWidget(self.barcode_input)
        layout.addWidget(QLabel("Price"))
        layout.addWidget(self.price_input)

        add_button = QPushButton("Add")
        add_button.clicked.connect(self.add_product)
        layout.addWidget(add_button)

        self.setLayout(layout)

    def add_product(self):
        name = self.name_input.text()
        barcode = self.barcode_input.text()
        price = self.price_input.text()

        # Add to the table
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        self.table.setItem(row_position, 0, QTableWidgetItem(name))
        self.table.setItem(row_position, 1, QTableWidgetItem(barcode))
        self.table.setItem(row_position, 2, QTableWidgetItem(price))

        # Add to the database
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            INSERT INTO products (name, barcode, unit_price) VALUES (?, ?, ?)
        ''', (name, barcode, price))
        connection.commit()
        connection.close()

        self.accept()

class EditProductWindow(QDialog):
    def __init__(self, table, row):
        super().__init__()
        self.setWindowTitle("Edit Product")
        self.setGeometry(100, 100, 300, 200)
        self.table = table
        self.row = row

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.barcode_input = QLineEdit()
        self.price_input = QLineEdit()

        layout.addWidget(QLabel("Name"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Barcode"))
        layout.addWidget(self.barcode_input)
        layout.addWidget(QLabel("Price"))
        layout.addWidget(self.price_input)

        self.load_product()

        edit_button = QPushButton("Edit")
        edit_button.clicked.connect(self.edit_product)
        layout.addWidget(edit_button)

        self.setLayout(layout)

    def load_product(self):
        self.name_input.setText(self.table.item(self.row, 0).text())
        self.barcode_input.setText(self.table.item(self.row, 1).text())
        self.price_input.setText(self.table.item(self.row, 2).text())

    def edit_product(self):
        name = self.name_input.text()
        barcode = self.barcode_input.text()
        price = self.price_input.text()

        # Update the table
        self.table.setItem(self.row, 0, QTableWidgetItem(name))
        self.table.setItem(self.row, 1, QTableWidgetItem(barcode))
        self.table.setItem(self.row, 2, QTableWidgetItem(price))

        # Update the database
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            UPDATE products SET name = ?, barcode = ?, unit_price = ? WHERE barcode = ?
        ''', (name, barcode, price, self.table.item(self.row, 1).text()))
        connection.commit()
        connection.close()

        self.accept()

class CashRegisterWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cash Register")
        self.setGeometry(100, 100, 600, 400)
        main_layout = QVBoxLayout()

        # First Frame
        first_frame = QFrame()
        first_frame.setFrameShape(QFrame.Shape.NoFrame)
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        first_title = QLabel("First Title")
        first_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        first_layout.addWidget(first_title)

        first_row_layout = QHBoxLayout()
        first_label = QLabel("Label:")
        first_input = QLineEdit()
        plus_button = QPushButton("+")
        minus_button = QPushButton("-")
        first_row_layout.addWidget(first_label)
        first_row_layout.addWidget(first_input)
        first_row_layout.addWidget(plus_button)
        first_row_layout.addWidget(minus_button)
        first_layout.addLayout(first_row_layout)

        # Second Frame
        second_frame = QFrame()
        second_frame.setFrameShape(QFrame.Shape.NoFrame)
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        second_title = QLabel("Second Title")
        second_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        second_layout.addWidget(second_title)

        second_row_layout1 = QHBoxLayout()
        second_label1 = QLabel("Label 1:")
        second_label2 = QLabel("Label 2:")
        second_row_layout1.addWidget(second_label1)
        second_row_layout1.addWidget(second_label2)
        second_layout.addLayout(second_row_layout1)

        second_row_layout2 = QHBoxLayout()
        second_input1 = QCalendarWidget()
        second_input2 = QCalendarWidget()
        second_row_layout2.addWidget(second_input1)
        second_row_layout2.addWidget(second_input2)
        second_layout.addLayout(second_row_layout2)

        second_button = QPushButton("Button")
        second_layout.addWidget(second_button)

        # Third Frame
        third_frame = QFrame()
        third_frame.setFrameShape(QFrame.Shape.NoFrame)
        third_layout = QVBoxLayout()
        third_frame.setLayout(third_layout)

        third_title = QLabel("Third Title")
        third_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        third_layout.addWidget(third_title)

        third_row_layout1 = QHBoxLayout()
        third_label1 = QLabel("Label 1:")
        third_label2 = QLabel("Label 2:")
        third_row_layout1.addWidget(third_label1)
        third_row_layout1.addWidget(third_label2)
        third_layout.addLayout(third_row_layout1)

        third_row_layout2 = QHBoxLayout()
        third_input1 = QCalendarWidget()
        third_input2 = QCalendarWidget()
        third_row_layout2.addWidget(third_input1)
        third_row_layout2.addWidget(third_input2)
        third_layout.addLayout(third_row_layout2)

        third_button = QPushButton("Button")
        third_layout.addWidget(third_button)

        # Fourth Frame
        fourth_frame = QFrame()
        fourth_frame.setFrameShape(QFrame.Shape.NoFrame)
        fourth_layout = QVBoxLayout()
        fourth_frame.setLayout(fourth_layout)

        fourth_title = QLabel("Fourth Title")
        fourth_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        fourth_layout.addWidget(fourth_title)

        fourth_button = QPushButton("Button")
        fourth_layout.addWidget(fourth_button)

        # Add frames to main layout
        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)
        main_layout.addWidget(third_frame)
        main_layout.addWidget(fourth_frame)

        self.setLayout(main_layout)

class SpravkiDocumentsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Spravki Documents")
        self.setGeometry(100, 100, 400, 300)
        layout = QVBoxLayout()
        label = QLabel("Spravki Documents Window")
        layout.addWidget(label)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        self.setLayout(layout)

class SpravkiClientsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Spravki Clients")
        self.setGeometry(100, 100, 400, 300)
        layout = QVBoxLayout()
        label = QLabel("Spravki Clients Window")
        layout.addWidget(label)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        self.setLayout(layout)

class SpravkiSalesWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Spravki Sales")
        self.setGeometry(100, 100, 400, 300)
        layout = QVBoxLayout()
        label = QLabel("Spravki Sales Window")
        layout.addWidget(label)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        self.setLayout(layout)

# class SettingsWindow(QDialog):
#     def __init__(self):
#         super().__init__()
#         self.setWindowTitle("Settings")
#         self.setGeometry(100, 100, 400, 300)
#         main_layout = QVBoxLayout()
#
#         # First Frame
#         first_frame = QFrame()
#         first_frame.setFrameShape(QFrame.Shape.NoFrame)
#         first_layout = QVBoxLayout()
#         first_frame.setLayout(first_layout)
#
#         first_title = QLabel("First Title")
#         first_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         first_layout.addWidget(first_title)
#
#         for i in range(4):
#             row_layout = QHBoxLayout()
#             label = QLabel(f"Label {i + 1}:")
#             input = QComboBox()
#             input.addItems([f"Option {j + 1}" for j in range(3)])
#             row_layout.addWidget(label)
#             row_layout.addWidget(input)
#             first_layout.addLayout(row_layout)
#
#         # Second Frame
#         second_frame = QFrame()
#         second_frame.setFrameShape(QFrame.Shape.NoFrame)
#         second_layout = QVBoxLayout()
#         second_frame.setLayout(second_layout)
#
#         second_title = QLabel("Second Title")
#         second_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         second_layout.addWidget(second_title)
#
#         for i in range(2):
#             row_layout = QHBoxLayout()
#             label = QLabel(f"Label {i + 1}:")
#             input = QLineEdit()
#             row_layout.addWidget(label)
#             row_layout.addWidget(input)
#             second_layout.addLayout(row_layout)
#
#         # Third Frame
#         third_frame = QFrame()
#         third_frame.setFrameShape(QFrame.Shape.NoFrame)
#         third_layout = QVBoxLayout()
#         third_frame.setLayout(third_layout)
#
#         third_title = QLabel("Third Title")
#         third_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
#         third_layout.addWidget(third_title)
#
#         for i in range(5):
#             row_layout = QHBoxLayout()
#             label = QLabel(f"Label {i + 1}:")
#             input = QComboBox()
#             input.addItems([f"Option {j + 1}" for j in range(3)])
#             row_layout.addWidget(label)
#             row_layout.addWidget(input)
#             third_layout.addLayout(row_layout)
#
#         # Add frames to main layout
#         main_layout.addWidget(first_frame)
#         main_layout.addWidget(second_frame)
#         main_layout.addWidget(third_frame)
#
#         self.setLayout(main_layout)

class SettingsWindow(QDialog):
    def __init__(self, main_window):
        super().__init__()
        self.setWindowTitle("Settings Window")
        self.setGeometry(100, 100, 600, 400)
        self.main_window = main_window


        main_layout = QVBoxLayout()

        # Create a QTabWidget
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # Add the first tab
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "General Settings")
        self.tab1_layout = QVBoxLayout()
        self.tab1.setLayout(self.tab1_layout)

        # First Frame
        first_frame = QFrame()
        first_frame.setFrameShape(QFrame.Shape.NoFrame)
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        first_title = QLabel("First Title")
        first_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        first_layout.addWidget(first_title)

        for i in range(4):
            row_layout = QHBoxLayout()
            label = QLabel(f"Label {i + 1}:")
            input = QComboBox()
            input.addItems([f"Option {j + 1}" for j in range(3)])
            row_layout.addWidget(label)
            row_layout.addWidget(input)
            first_layout.addLayout(row_layout)

        # Second Frame
        second_frame = QFrame()
        second_frame.setFrameShape(QFrame.Shape.NoFrame)
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        second_title = QLabel("Second Title")
        second_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        second_layout.addWidget(second_title)

        self.operator_name_input = QLineEdit()
        self.operator_number_input = QLineEdit()
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.save_settings)

        second_layout.addWidget(QLabel("Operator Name"))
        second_layout.addWidget(self.operator_name_input)
        second_layout.addWidget(QLabel("Operator Number"))
        second_layout.addWidget(self.operator_number_input)
        second_layout.addWidget(save_button)

        # Third Frame
        third_frame = QFrame()
        third_frame.setFrameShape(QFrame.Shape.NoFrame)
        third_layout = QVBoxLayout()
        third_frame.setLayout(third_layout)

        third_title = QLabel("Third Title")
        third_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        third_layout.addWidget(third_title)

        for i in range(5):
            row_layout = QHBoxLayout()
            label = QLabel(f"Label {i + 1}:")
            input = QComboBox()
            input.addItems([f"Option {j + 1}" for j in range(3)])
            row_layout.addWidget(label)
            row_layout.addWidget(input)
            third_layout.addLayout(row_layout)

        # Add frames to the first tab layout
        self.tab1_layout.addWidget(first_frame)
        self.tab1_layout.addWidget(second_frame)
        self.tab1_layout.addWidget(third_frame)

        # Add the second tab
        self.tab2 = QWidget()
        self.tabs.addTab(self.tab2, "Add Firm")
        self.tab2_layout = QVBoxLayout()
        self.tab2.setLayout(self.tab2_layout)

        # Add labels and inputs to the second tab
        self.firm_name_label = QLabel("Firm Name:")
        self.firm_name_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_name_label)
        self.tab2_layout.addWidget(self.firm_name_input)

        self.firm_eik_label = QLabel("EIK:")
        self.firm_eik_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_eik_label)
        self.tab2_layout.addWidget(self.firm_eik_input)

        self.firm_dds_label = QLabel("DDS:")
        self.firm_dds_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_dds_label)
        self.tab2_layout.addWidget(self.firm_dds_input)

        self.firm_address_label = QLabel("Address:")
        self.firm_address_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_address_label)
        self.tab2_layout.addWidget(self.firm_address_input)

        self.firm_mol_label = QLabel("MOL:")
        self.firm_mol_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_mol_label)
        self.tab2_layout.addWidget(self.firm_mol_input)

        self.firm_phone_label = QLabel("Phone:")
        self.firm_phone_input = QLineEdit()
        self.tab2_layout.addWidget(self.firm_phone_label)
        self.tab2_layout.addWidget(self.firm_phone_input)

        self.add_firm_button = QPushButton("Add Firm")
        self.add_firm_button.clicked.connect(self.add_firm)
        self.tab2_layout.addWidget(self.add_firm_button)

        self.setLayout(main_layout)
    #
    # def save_settings(self):
    #     # Implement the code to save the general settings
    #     setting1 = self.setting1_input.text()
    #     setting2 = self.setting2_input.text()
    #     print(f"Saved General Settings: {setting1}, {setting2}")

    def save_additional_settings(self):
        # Implement the code to save the additional settings
        additional_setting1 = self.additional_setting1_input.text()
        additional_setting2 = self.additional_setting2_input.text()
        print(f"Saved Additional Settings: {additional_setting1}, {additional_setting2}")

    def save_settings(self):
        operator_name = self.operator_name_input.text()
        operator_number = self.operator_number_input.text()
        if operator_name and operator_number:
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('''
                INSERT INTO settings (operator_name, operator_number)
                VALUES (?, ?)
            ''', (operator_name, operator_number))
            connection.commit()
            connection.close()
            self.main_window.update_operator_label(f"{operator_name} - {operator_number}")
            self.accept()

    def add_firm(self):
        # Implement the code to add a firm
        firm_name = self.firm_name_input.text()
        firm_eik = self.firm_eik_input.text()
        firm_dds = self.firm_dds_input.text()
        firm_address = self.firm_address_input.text()
        firm_mol = self.firm_mol_input.text()
        firm_phone = self.firm_phone_input.text()

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            INSERT INTO firm (name, eik, dds, address, mol, phone)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (firm_name, firm_eik, firm_dds, firm_address, firm_mol, firm_phone))
        connection.commit()
        connection.close()

        print(f"Added Firm: {firm_name}, {firm_eik}, {firm_dds}, {firm_address}, {firm_mol}, {firm_phone}")
        self.firm_name_input.clear()
        self.firm_eik_input.clear()
        self.firm_dds_input.clear()
        self.firm_address_input.clear()
        self.firm_mol_input.clear()
        self.firm_phone_input.clear()

class HelpWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Help")
        self.setGeometry(100, 100, 400, 300)
        layout = QVBoxLayout()
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setText("This is the help text for the program. Here you can provide information about how to use the application, tips, and any other relevant information.")
        layout.addWidget(help_text)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        self.setLayout(layout)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Main Window with Three Horizontal Frames and Menubar")
        self.setWindowIcon(QIcon('vladpos_logo.png'))
        self.setGeometry(100, 100, 800, 600)
        self.showMaximized()

        # Set up the central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Create frames
        frame1 = QFrame()
        frame1.setFrameShape(QFrame.Shape.NoFrame)
        frame1.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)

        frame2 = QFrame()
        frame2.setFrameShape(QFrame.Shape.NoFrame)
        frame2.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)

        frame3 = QFrame()
        frame3.setFrameShape(QFrame.Shape.NoFrame)
        frame3.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)

        # Add frames to layout
        main_layout.addWidget(frame1, 1)
        main_layout.addWidget(frame2, 5)
        main_layout.addWidget(frame3, 2)

        # Create menubar
        menubar = self.menuBar()

        # Create Cash Register menu
        cash_register_menu = QMenu("Cash Register", self)
        cash_register_menu.setIcon(QIcon('receipt.png'))
        cash_register_action = QAction(QIcon('receipt.png'), 'Cash Register Action', self)
        cash_register_action.triggered.connect(self.open_cash_register_window)
        cash_register_menu.addAction(cash_register_action)
        menubar.addMenu(cash_register_menu)

        # Create Spravki menu
        spravki_menu = QMenu("Spravki", self)
        spravki_menu.setIcon(QIcon('paper.png'))

        # Spravki Documents submenu
        spravki_documents_action = QAction('Spravki Documents', self)
        spravki_documents_action.triggered.connect(self.open_spravki_documents_window)
        spravki_menu.addAction(spravki_documents_action)

        # Spravki Clients submenu
        spravki_clients_action = QAction('Spravki Clients', self)
        spravki_clients_action.triggered.connect(self.open_spravki_clients_window)
        spravki_menu.addAction(spravki_clients_action)

        # Spravki Sales submenu
        spravki_sales_action = QAction('Spravki Sales', self)
        spravki_sales_action.triggered.connect(self.open_spravki_sales_window)
        spravki_menu.addAction(spravki_sales_action)

        menubar.addMenu(spravki_menu)

        # Create New Menu
        new_menu = QMenu("New Menu", self)
        new_menu.setIcon(QIcon('products.png'))
        new_action = QAction(QIcon('new_action_icon.png'), 'Open New Window', self)
        new_action.triggered.connect(self.open_new_window)
        new_menu.addAction(new_action)

        menubar.addMenu(new_menu)

        # Add the following code in the menubar section to create the Customers menu
        customers_menu = QMenu("Customers", self)
        customers_menu.setIcon(QIcon('user-2.png'))
        customers_action = QAction(QIcon('customer_action_icon.png'), 'Open Customers Window', self)
        customers_action.triggered.connect(self.open_customers_window)
        customers_menu.addAction(customers_action)
        menubar.addMenu(customers_menu)

        # Add the following code in the menubar section to create the Documents menu
        documents_menu = QMenu("Documents", self)
        documents_menu.setIcon(QIcon('paper.png'))
        documents_action = QAction(QIcon('document_action_icon.png'), 'Open Documents Window', self)
        documents_action.triggered.connect(self.open_documents_window)
        documents_menu.addAction(documents_action)
        menubar.addMenu(documents_menu)

        # Add the following code in the menubar section to create the Invoices menu
        invoices_menu = QMenu("Invoices", self)
        invoices_menu.setIcon(QIcon('invoice.png'))
        invoices_action = QAction(QIcon('invoice.png'), 'Open Invoices Window', self)
        invoices_action.triggered.connect(self.open_invoices_window)
        invoices_menu.addAction(invoices_action)
        menubar.addMenu(invoices_menu)

        # Create Settings menu
        settings_menu = QMenu("Settings", self)
        settings_menu.setIcon(QIcon('setting.png'))
        settings_action = QAction(QIcon('setting.png'), 'Settings Action', self)
        settings_action.triggered.connect(self.open_settings_window)
        settings_menu.addAction(settings_action)
        menubar.addMenu(settings_menu)

        # Create Help menu
        help_menu = QMenu("Help", self)
        help_menu.setIcon(QIcon('question.png'))
        help_action = QAction(QIcon('help.png'), 'Help', self)
        help_action.triggered.connect(self.open_help_window)
        help_menu.addAction(help_action)
        menubar.addMenu(help_menu)

        # Frame 1 Contents
        frame1_layout = QHBoxLayout()
        frame1.setLayout(frame1_layout)

        self.input1 = QLineEdit()
        self.input1.setPlaceholderText("Input 1")
        self.input1.textChanged.connect(self.suggestions)
        frame1_layout.addWidget(self.input1, 3)  # Biggest

        self.suggestions_list = QListWidget()
        self.suggestions_list.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.suggestions_list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.suggestions_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.suggestions_list.setStyleSheet("QListWidget { background-color: white; border: 1px solid gray; }")
        self.suggestions_list.hide()
        frame1_layout.addWidget(self.suggestions_list)
        self.suggestions_list.itemClicked.connect(self.select_suggestion)

        self.timer = QTimer()
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.hide_suggestions)

        self.input2 = QLineEdit()
        self.input2.setPlaceholderText("Input 2")
        frame1_layout.addWidget(self.input2, 1)  # Small

        self.input3 = QLineEdit()
        self.input3.setPlaceholderText("Input 3")
        frame1_layout.addWidget(self.input3, 1)  # Like the second

        self.popup_menu1 = QComboBox()
        self.popup_menu1.addItems(["Option 1", "Option 2", "Option 3", "Option 4"])
        frame1_layout.addWidget(self.popup_menu1, 2)  # A little bit more bigger

        self.add_button = QPushButton("Add")
        self.add_button.setIcon(QIcon('ecommerce.png'))
        frame1_layout.addWidget(self.add_button, 2)  # Bigger

        # Frame 2 Contents
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Input 1", "Input 2", "Input 3", "Popup Menu", "Multiplication", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        frame2_layout = QVBoxLayout()
        frame2.setLayout(frame2_layout)
        frame2_layout.addWidget(self.table)

        # Frame 3 Contents
        frame3_layout = QVBoxLayout()
        frame3.setLayout(frame3_layout)

        sub_frame1 = QFrame()
        sub_frame1.setFrameShape(QFrame.Shape.NoFrame)
        sub_frame1_layout = QVBoxLayout()
        sub_frame1.setLayout(sub_frame1_layout)
        image_label = QLabel()
        image_label.setPixmap(QPixmap('user.png'))
        image_label.setScaledContents(True)  # Scale the image to fit the label
        image_label.setMaximumSize(100,100)  # Set maximum size for the image label
        sub_frame1_layout.addWidget(image_label, alignment=Qt.AlignmentFlag.AlignCenter)
        self.image_text_label = QLabel("Image Label")
        self.image_text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Center the text under the image
        sub_frame1_layout.addWidget(self.image_text_label)

        sub_frame2 = QFrame()
        sub_frame2.setFrameShape(QFrame.Shape.NoFrame)
        sub_frame2_layout = QVBoxLayout()
        sub_frame2.setLayout(sub_frame2_layout)
        self.popup_menu2 = QComboBox()
        self.popup_menu2.addItems(["Option 1", "Option 2", "Option 3", "Option 4"])
        sub_frame2_layout.addWidget(self.popup_menu2)
        button1 = QPushButton("Button 1")
        sub_frame2_layout.addWidget(button1)
        button2 = QPushButton("Button 2")
        sub_frame2_layout.addWidget(button2)

        sub_frame3 = QFrame()
        sub_frame3.setFrameShape(QFrame.Shape.NoFrame)
        sub_frame3_layout = QVBoxLayout()
        sub_frame3.setLayout(sub_frame3_layout)
        self.input4 = QLineEdit()
        self.input4.setPlaceholderText("Input 4")
        sub_frame3_layout.addWidget(self.input4)
        button3 = QPushButton("Button 3")
        sub_frame3_layout.addWidget(button3)
        button4 = QPushButton("Button 4")
        sub_frame3_layout.addWidget(button4)

        sub_frame4 = QFrame()
        sub_frame4.setFrameShape(QFrame.Shape.NoFrame)
        sub_frame4_layout = QVBoxLayout()
        sub_frame4.setLayout(sub_frame4_layout)
        self.input5 = QLineEdit()
        self.input5.setPlaceholderText("Input 5")
        sub_frame4_layout.addWidget(self.input5)
        self.input6 = QLineEdit()
        self.input6.setPlaceholderText("Input 6")
        sub_frame4_layout.addWidget(self.input6)
        button5 = QPushButton("Button 5")
        sub_frame4_layout.addWidget(button5)
        # button6 = QPushButton("Button 6")
        # sub_frame4_layout.addWidget(button6)
        self.input5.textChanged.connect(self.update_change)
        button1.clicked.connect(self.save_to_database)


        frame3_sub_layout = QHBoxLayout()
        frame3_layout.addLayout(frame3_sub_layout)
        frame3_sub_layout.addWidget(sub_frame1)
        frame3_sub_layout.addWidget(sub_frame2)
        frame3_sub_layout.addWidget(sub_frame3)
        frame3_sub_layout.addWidget(sub_frame4)

        # Set equal stretch for sub-frames
        frame3_sub_layout.setStretch(0, 1)
        frame3_sub_layout.setStretch(1, 1)
        frame3_sub_layout.setStretch(2, 1)
        frame3_sub_layout.setStretch(3, 1)

        # Connect the add button to the function
        self.add_button.clicked.connect(self.add_to_table)

    def suggestions(self):
        text = self.input1.text()
        if text:
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('''
                SELECT name, barcode FROM products
                WHERE name LIKE ? OR barcode LIKE ?
            ''', ('%' + text + '%', '%' + text + '%'))
            suggestions = cursor.fetchall()
            connection.close()

            self.suggestions_list.clear()
            for suggestion in suggestions:
                self.suggestions_list.addItem(f"{suggestion[0]} ({suggestion[1]})")

            self.suggestions_list.show()
        else:
            self.suggestions_list.hide()

    def hide_suggestions(self):
        self.suggestions_list.hide()

    def update_operator_label(self, text):
        self.image_text_label.setText(text)

    def select_suggestion(self, item):
        text = item.text()
        name = text.split(' (')[0]
        self.input1.setText(name)
        self.suggestions_list.hide()

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            SELECT unit_price FROM products
            WHERE name = ?
        ''', (name,))
        price = cursor.fetchone()
        connection.close()

        if price and price[0] != 0:
            self.input3.setText(str(price[0]))

    def update_totals(self):
        total_sum = 0
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 4)
            if item:
                total_sum += float(item.text())

        self.input4.setText(str(total_sum))
        self.input5.setText(str(total_sum))
        self.update_change()

    def update_change(self):
        try:
            total_sum = float(self.input4.text())
            paid_amount = float(self.input5.text())
            change = paid_amount - total_sum
        except ValueError:
            change = 0

        self.input6.setText(str(change))

    def save_to_database(self):
        products = []
        for row in range(self.table.rowCount()):
            product = {
                "name": self.table.item(row, 0).text(),
                # "barcode": self.table.item(row, 1).text(),
                "quantity": float(self.table.item(row, 1).text()),
                "unit_price": float(self.table.item(row, 2).text()),
                "total_price": float(self.table.item(row, 4).text())
            }
            products.append(product)

        total_amount = float(self.input4.text())
        amount_paid = float(self.input5.text())
        change = float(self.input6.text())
        payment_type = self.popup_menu2.currentText()
        cash_register_numbers = ""

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            INSERT INTO documents (products, total_amount, amount_paid, change, payment_type, cash_register_numbers)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (json.dumps(products), total_amount, amount_paid, change, payment_type, cash_register_numbers))
        connection.commit()
        connection.close()

        self.table.setRowCount(0)
        self.input4.clear()
        self.input5.clear()
        self.input6.clear()

        msg = QMessageBox()
        # msg.setIcon(QIcon('add-to-basket.png'))
        msg.setText("Data saved successfully!")
        msg.setWindowTitle("Success")
        msg.exec()

    def add_to_table(self):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        input1_text = self.input1.text()
        input2_text = self.input2.text()
        input3_text = self.input3.text()
        popup_text = self.popup_menu1.currentText()

        self.table.setItem(row_position, 0, QTableWidgetItem(input1_text))
        self.table.setItem(row_position, 1, QTableWidgetItem(input2_text))
        self.table.setItem(row_position, 2, QTableWidgetItem(input3_text))
        self.table.setItem(row_position, 3, QTableWidgetItem(popup_text))

        try:
            multiplication_result = float(input2_text) * float(input3_text)
        except ValueError:
            multiplication_result = 0
        self.table.setItem(row_position, 4, QTableWidgetItem(str(multiplication_result)))

        delete_button = QPushButton()
        delete_button.setIcon(QIcon('trash_icon.png'))
        delete_button.clicked.connect(lambda: self.remove_row(row_position))
        self.table.setCellWidget(row_position, 5, delete_button)

        self.input1.clear()
        self.input2.clear()
        self.input3.clear()
        self.popup_menu1.setCurrentIndex(0)

        self.update_totals()

    def remove_row(self, row):
        self.table.removeRow(row)

    def open_cash_register_window(self):
        self.cash_register_window = CashRegisterWindow()
        self.cash_register_window.exec()

    def open_spravki_documents_window(self):
        self.spravki_documents_window = SpravkiDocumentsWindow()
        self.spravki_documents_window.exec()

    def open_spravki_clients_window(self):
        self.spravki_clients_window = SpravkiClientsWindow()
        self.spravki_clients_window.exec()

    def open_spravki_sales_window(self):
        self.spravki_sales_window = SpravkiSalesWindow()
        self.spravki_sales_window.exec()

    def open_settings_window(self):
        self.settings_window = SettingsWindow(self)
        self.settings_window.exec()

    def open_help_window(self):
        self.help_window = HelpWindow()
        self.help_window.exec()

    def open_new_window(self):
        self.new_window = ProductsWindow()
        self.new_window.exec()

    def open_customers_window(self):
        self.customers_window = CustomersWindow()
        self.customers_window.exec()

    def open_documents_window(self):
        self.documents_window = DocumentsWindow()
        self.documents_window.exec()

    def open_invoices_window(self):
        self.invoices_window = InvoicesWindow()
        self.invoices_window.exec()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
