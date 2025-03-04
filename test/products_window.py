import sys
import os
import sqlite3
import json
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QIcon, QAction, QPixmap
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFrame, QVBoxLayout, QHBoxLayout, QWidget, QMenu,
    QSizePolicy, QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialog, QDialogButtonBox, QCalendarWidget, QTextEdit, QListWidget,
    QFileDialog, QAbstractItemView, QMessageBox, QTabWidget
)

DB_DIR = "database"
DB_PATH = "database/pos_system.db"

class ProductsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Продукти")
        self.setGeometry(100, 100, 600, 400)

        main_layout = QVBoxLayout()

        # First Frame
        first_frame = QFrame()
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Име", "Баркод", "Ед. цена"])
        first_layout.addWidget(self.table)

        # Load products into the table
        self.load_products()

        # Second Frame
        second_frame = QFrame()
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        buttons_layout = QHBoxLayout()
        self.add_button = QPushButton("Нов")
        self.add_button.clicked.connect(self.open_add_product_window)
        buttons_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Редактирай")
        self.edit_button.clicked.connect(self.open_edit_product_window)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Изтрий")
        self.delete_button.clicked.connect(self.delete_product)
        buttons_layout.addWidget(self.delete_button)

        self.refresh_button = QPushButton("Обнови")
        self.refresh_button.clicked.connect(self.load_products)
        buttons_layout.addWidget(self.refresh_button)

        self.export_button = QPushButton("Експортиране в XLSX")
        self.export_button.clicked.connect(self.export_to_xlsx)
        buttons_layout.addWidget(self.export_button)

        self.import_button = QPushButton("Импортиране от XLSX")
        self.import_button.clicked.connect(self.import_from_xlsx)
        buttons_layout.addWidget(self.import_button)

        second_layout.addLayout(buttons_layout)

        combo_layout = QHBoxLayout()
        self.combo1 = QComboBox()
        self.combo2 = QComboBox()
        combo_layout.addWidget(self.combo1)
        combo_layout.addWidget(self.combo2)
        second_layout.addLayout(combo_layout)

        # Third Frame

        # Add these lines in the __init__ method to create filter inputs
        self.filter_frame = QFrame()
        self.filter_layout = QHBoxLayout()
        self.filter_frame.setLayout(self.filter_layout)
        self.filter_name_input = QLineEdit()
        self.filter_name_input.setPlaceholderText("Филтър по име")
        self.filter_name_input.textChanged.connect(self.filter_products)
        self.filter_layout.addWidget(self.filter_name_input)
        self.filter_barcode_input = QLineEdit()
        self.filter_barcode_input.setPlaceholderText("Филтър по баркод")
        self.filter_barcode_input.textChanged.connect(self.filter_products)
        self.filter_layout.addWidget(self.filter_barcode_input)
        first_layout.addWidget(self.filter_frame)


        # Add frames to main layout
        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)
        main_layout.addWidget(self.filter_frame)



        self.setLayout(main_layout)

    def load_products(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT name, barcode, unit_price FROM products')
        products = cursor.fetchall()

        self.table.setRowCount(len(products))
        for row_idx, product in enumerate(products):
            self.table.setItem(row_idx, 0, QTableWidgetItem(product[0]))
            self.table.setItem(row_idx, 1, QTableWidgetItem(product[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(product[2])))

        connection.close()

    def open_add_product_window(self):
        self.add_product_window = AddProductWindow(self.table)
        self.add_product_window.exec()

    def open_edit_product_window(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.edit_product_window = EditProductWindow(self.table, current_row)
            self.edit_product_window.exec()

    def delete_product(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            barcode = self.table.item(current_row, 1).text()
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM products WHERE barcode = ?', (barcode,))
            connection.commit()
            connection.close()
            self.table.removeRow(current_row)
            self.load_products()

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Продукти"

            headers = ["Име", "Баркод", "Ед. цена"]
            sheet.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            workbook.save(file_path)

    def import_from_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            self.table.setRowCount(0)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for column, value in enumerate(row):
                    self.table.setItem(row_position, column, QTableWidgetItem(str(value)))

            # Optionally, update the database with the imported data
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM products')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cursor.execute('''
                    INSERT INTO products (name, barcode, unit_price) VALUES (?, ?, ?)
                ''', row)
            connection.commit()
            connection.close()

    # Add this method to filter the products table
    def filter_products(self):
        name_filter = self.filter_name_input.text().lower()
        barcode_filter = self.filter_barcode_input.text().lower()

        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            barcode_item = self.table.item(row, 1)
            name_match = name_filter in name_item.text().lower() if name_item else False
            barcode_match = barcode_filter in barcode_item.text().lower() if barcode_item else False
            self.table.setRowHidden(row, not (name_match and barcode_match))

class AddProductWindow(QDialog):
    def __init__(self, table):
        super().__init__()
        self.setWindowTitle("Нов")
        self.setGeometry(100, 100, 300, 200)
        self.table = table

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.barcode_input = QLineEdit()
        self.price_input = QLineEdit()

        layout.addWidget(QLabel("Име"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Баркод"))
        layout.addWidget(self.barcode_input)
        layout.addWidget(QLabel("Ед. цена"))
        layout.addWidget(self.price_input)

        add_button = QPushButton("Добави")
        add_button.clicked.connect(self.add_product)
        layout.addWidget(add_button)

        self.setLayout(layout)

    def add_product(self):
        name = self.name_input.text()
        barcode = self.barcode_input.text()
        price = self.price_input.text()

        # Add to the table
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        self.table.setItem(row_position, 0, QTableWidgetItem(name))
        self.table.setItem(row_position, 1, QTableWidgetItem(barcode))
        self.table.setItem(row_position, 2, QTableWidgetItem(price))

        # Add to the database
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            INSERT INTO products (name, barcode, unit_price) VALUES (?, ?, ?)
        ''', (name, barcode, price))
        connection.commit()
        connection.close()

        self.accept()

class EditProductWindow(QDialog):
    def __init__(self, table, row):
        super().__init__()
        self.setWindowTitle("Редактиране")
        self.setGeometry(100, 100, 300, 200)
        self.table = table
        self.row = row

        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.barcode_input = QLineEdit()
        self.price_input = QLineEdit()

        layout.addWidget(QLabel("Име"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("Баркод"))
        layout.addWidget(self.barcode_input)
        layout.addWidget(QLabel("Цена"))
        layout.addWidget(self.price_input)

        self.load_product()

        edit_button = QPushButton("Запази")
        edit_button.clicked.connect(self.edit_product)
        layout.addWidget(edit_button)

        self.setLayout(layout)

    def load_product(self):
        self.name_input.setText(self.table.item(self.row, 0).text())
        self.barcode_input.setText(self.table.item(self.row, 1).text())
        self.price_input.setText(self.table.item(self.row, 2).text())

    def edit_product(self):
        name = self.name_input.text()
        barcode = self.barcode_input.text()
        price = self.price_input.text()

        # Update the table
        self.table.setItem(self.row, 0, QTableWidgetItem(name))
        self.table.setItem(self.row, 1, QTableWidgetItem(barcode))
        self.table.setItem(self.row, 2, QTableWidgetItem(price))

        # Update the database
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            UPDATE products SET name = ?, barcode = ?, unit_price = ? WHERE barcode = ?
        ''', (name, barcode, price, self.table.item(self.row, 1).text()))
        connection.commit()
        connection.close()

        self.accept()