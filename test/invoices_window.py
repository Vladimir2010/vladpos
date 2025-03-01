import sqlite3
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QPushButton,
    QComboBox, QLineEdit, QFrame, QFileDialog, QDateEdit, QCalendarWidget, QAbstractItemView
)
from PyQt6.QtCore import Qt, QDate
from openpyxl import Workbook, load_workbook

DB_PATH = "database/pos_system.db"


class InvoicesWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Invoices Window")
        self.setGeometry(100, 100, 800, 600)

        main_layout = QVBoxLayout()

        first_frame = QFrame()
        first_layout = QVBoxLayout()
        first_frame.setLayout(first_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Number", "Customer Ref No", "Selling Company Ref No", "Date", "Products",
            "Total Amount", "Amount Paid", "Change", "Payment Type"
        ])
        first_layout.addWidget(self.table)

        self.load_invoices()

        second_frame = QFrame()
        second_layout = QVBoxLayout()
        second_frame.setLayout(second_layout)

        buttons_layout = QHBoxLayout()

        self.add_button = QPushButton("Add New Invoice")
        self.add_button.clicked.connect(self.open_add_invoice_window)
        buttons_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Edit Invoice")
        self.edit_button.clicked.connect(self.open_edit_invoice_window)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Delete Invoice")
        self.delete_button.clicked.connect(self.delete_invoice)
        buttons_layout.addWidget(self.delete_button)

        self.refresh_button = QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.load_invoices)
        buttons_layout.addWidget(self.refresh_button)

        self.export_button = QPushButton("Export to XLSX")
        self.export_button.clicked.connect(self.export_to_xlsx)
        buttons_layout.addWidget(self.export_button)

        self.import_button = QPushButton("Import from XLSX")
        self.import_button.clicked.connect(self.import_from_xlsx)
        buttons_layout.addWidget(self.import_button)

        second_layout.addLayout(buttons_layout)

        main_layout.addWidget(first_frame)
        main_layout.addWidget(second_frame)

        self.setLayout(main_layout)

    def load_invoices(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            SELECT number, customer_ref_no, selling_company_ref_no, date, products,
                   total_amount, amount_paid, change, payment_type
            FROM invoices
        ''')
        invoices = cursor.fetchall()

        self.table.setRowCount(len(invoices))
        for row_idx, invoice in enumerate(invoices):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(invoice[0]).zfill(10)))
            self.table.setItem(row_idx, 1, QTableWidgetItem(invoice[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem(invoice[2]))
            self.table.setItem(row_idx, 3, QTableWidgetItem(invoice[3]))
            self.table.setItem(row_idx, 4, QTableWidgetItem(invoice[4]))
            self.table.setItem(row_idx, 5, QTableWidgetItem(str(invoice[5])))
            self.table.setItem(row_idx, 6, QTableWidgetItem(str(invoice[6])))
            self.table.setItem(row_idx, 7, QTableWidgetItem(str(invoice[7])))
            self.table.setItem(row_idx, 8, QTableWidgetItem(invoice[8]))

        connection.close()

    def open_add_invoice_window(self):
        self.add_invoice_window = AddInvoiceWindow(self.table)
        self.add_invoice_window.exec()

    def open_edit_invoice_window(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.edit_invoice_window = EditInvoiceWindow(self.table, current_row)
            self.edit_invoice_window.exec()

    def delete_invoice(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            invoice_id = self.table.item(current_row, 0).text()
            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM invoices WHERE number = ?', (invoice_id,))
            connection.commit()
            connection.close()
            self.table.removeRow(current_row)
            self.load_invoices()

    def export_to_xlsx(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"

            headers = ["Number", "Customer Ref No", "Selling Company Ref No", "Date", "Products",
                       "Total Amount", "Amount Paid", "Change", "Payment Type"]
            sheet.append(headers)

            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            workbook.save(file_path)

    def import_from_xlsx(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            self.table.setRowCount(0)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for column, value in enumerate(row):
                    self.table.setItem(row_position, column, QTableWidgetItem(str(value)))

            connection = sqlite3.connect(DB_PATH)
            cursor = connection.cursor()
            cursor.execute('DELETE FROM invoices')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cursor.execute('''
                    INSERT INTO invoices (number, customer_ref_no, selling_company_ref_no, date, products, 
                                          total_amount, amount_paid, change, payment_type) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', row)
            connection.commit()
            connection.close()

# class AddInvoiceWindow(QDialog):
#     def __init__(self, table):
#         super().__init__()
#         self.setWindowTitle("Add Invoice")
#         self.setGeometry(100, 100, 400, 500)
#         self.table = table
#
#         layout = QVBoxLayout()
#
#         self.number_input = QLineEdit()
#         self.number_input.setReadOnly(True)
#         self.number_input.setText(self.generate_invoice_number())
#         self.customer_ref_no_input = QLineEdit()
#         self.customer_ref_no_input.setReadOnly(True)
#         self.customer_ref_no_input.mouseDoubleClickEvent = self.open_customers_window
#         self.selling_company_ref_no_input = QLineEdit()
#         self.selling_company_ref_no_input.setReadOnly(True)
#         self.selling_company_ref_no_input.mouseDoubleClickEvent = self.open_firm_window
#         self.date_input = QDateEdit(calendarPopup=True)
#         self.date_input.setDate(QDate.currentDate())
#         self.products_input = QLineEdit()
#         self.products_input.setReadOnly(True)
#         self.products_input.mouseDoubleClickEvent = self.open_products_window
#         self.total_amount_input = QLineEdit()
#         self.amount_paid_input = QLineEdit()
#         self.change_input = QLineEdit()
#         self.payment_type_input = QLineEdit()
#
#         layout.addWidget(QLabel("Number"))
#         layout.addWidget(self.number_input)
#         layout.addWidget(QLabel("Customer Ref No"))
#         layout.addWidget(self.customer_ref_no_input)
#         layout.addWidget(QLabel("Selling Company Ref No"))
#         layout.addWidget(self.selling_company_ref_no_input)
#         layout.addWidget(QLabel("Date"))
#         layout.addWidget(self.date_input)
#         layout.addWidget(QLabel("Products"))
#         layout.addWidget(self.products_input)
#         layout.addWidget(QLabel("Total Amount"))
#         layout.addWidget(self.total_amount_input)
#         layout.addWidget(QLabel("Amount Paid"))
#         layout.addWidget(self.amount_paid_input)
#         layout.addWidget(QLabel("Change"))
#         layout.addWidget(self.change_input)
#         layout.addWidget(QLabel("Payment Type"))
#         layout.addWidget(self.payment_type_input)
#
#         add_button = QPushButton("Add")
#         add_button.clicked.connect(self.add_invoice)
#         layout.addWidget(add_button)
#
#         self.setLayout(layout)
#
#     def open_customers_window(self, event):
#         self.customers_window = CustomersWindow(self)
#         self.customers_window.exec()
#
#     def open_firm_window(self, event):
#         self.firm_window = FirmWindow(self)
#         self.firm_window.exec()
#
#     def open_products_window(self, event):
#         self.products_window = ProductsWindow(self)
#         self.products_window.exec()
#
#     def generate_invoice_number(self):
#         connection = sqlite3.connect(DB_PATH)
#         cursor = connection.cursor()
#         cursor.execute('SELECT MAX(number) FROM invoices')
#         result = cursor.fetchone()
#         connection.close()
#         if result[0] is not None:
#             return str(int(result[0]) + 1).zfill(10)
#         else:
#             return "0000000001"
#
#     def add_invoice(self):
#         number = self.number_input.text()
#         customer_ref_no = self.customer_ref_no_input.text()
#         selling_company_ref_no = self.selling_company_ref_no_input.text()
#         date = self.date_input.date().toString("yyyy-MM-dd")
#         products = self.products_input.text()
#         total_amount = self.total_amount_input.text()
#         amount_paid = self.amount_paid_input.text()
#         change = self.change_input.text()
#         payment_type = self.payment_type_input.text()
#
#         row_position = self.table.rowCount()
#         self.table.insertRow(row_position)
#         self.table.setItem(row_position, 0, QTableWidgetItem(number))
#         self.table.setItem(row_position, 1, QTableWidgetItem(customer_ref_no))
#         self.table.setItem(row_position, 2, QTableWidgetItem(selling_company_ref_no))
#         self.table.setItem(row_position, 3, QTableWidgetItem(date))
#         self.table.setItem(row_position, 4, QTableWidgetItem(products))
#         self.table.setItem(row_position, 5, QTableWidgetItem(total_amount))
#         self.table.setItem(row_position, 6, QTableWidgetItem(amount_paid))
#         self.table.setItem(row_position, 7, QTableWidgetItem(change))
#         self.table.setItem(row_position, 8, QTableWidgetItem(payment_type))
#
#         connection = sqlite3.connect(DB_PATH)
#         cursor = connection.cursor()
#         cursor.execute('''
#             INSERT INTO invoices (number, customer_ref_no, selling_company_ref_no, date, products,
#                                   total_amount, amount_paid, change, payment_type)
#             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
#         ''', (number, customer_ref_no, selling_company_ref_no, date, products, total_amount, amount_paid, change, payment_type))
#         connection.commit()
#         connection.close()
#
#         self.accept()

# In the AddInvoiceWindow class, modify the __init__ method as follows:
class AddInvoiceWindow(QDialog):
    def __init__(self, table):
        super().__init__()
        self.setWindowTitle("Add Invoice")
        self.setGeometry(100, 100, 400, 500)
        self.table = table

        layout = QVBoxLayout()

        self.number_input = QLineEdit()
        self.number_input.setReadOnly(True)
        self.number_input.setText(self.generate_invoice_number())
        self.customer_ref_no_input = QLineEdit()
        self.customer_ref_no_input.setReadOnly(True)
        self.customer_ref_no_input.mouseDoubleClickEvent = self.open_customers_window
        self.selling_company_ref_no_input = QLineEdit()
        self.selling_company_ref_no_input.setReadOnly(True)
        self.selling_company_ref_no_input.mouseDoubleClickEvent = self.open_firm_window
        self.date_input = QDateEdit(calendarPopup=True)
        self.date_input.setDate(QDate.currentDate())
        self.products_input = QLineEdit()
        self.products_input.setReadOnly(True)
        self.products_input.mouseDoubleClickEvent = self.open_products_window
        self.total_amount_input = QLineEdit()
        self.total_amount_input.setReadOnly(True)
        self.amount_paid_input = QLineEdit()
        self.amount_paid_input.textChanged.connect(self.update_change)
        self.change_input = QLineEdit()
        self.change_input.setReadOnly(True)
        self.payment_type_input = QComboBox()
        self.payment_type_input.addItems(["Cash", "Credit Card", "Debit Card", "Online Payment"])

        layout.addWidget(QLabel("Number"))
        layout.addWidget(self.number_input)
        layout.addWidget(QLabel("Customer Ref No"))
        layout.addWidget(self.customer_ref_no_input)
        layout.addWidget(QLabel("Selling Company Ref No"))
        layout.addWidget(self.selling_company_ref_no_input)
        layout.addWidget(QLabel("Date"))
        layout.addWidget(self.date_input)
        layout.addWidget(QLabel("Products"))
        layout.addWidget(self.products_input)
        layout.addWidget(QLabel("Total Amount"))
        layout.addWidget(self.total_amount_input)
        layout.addWidget(QLabel("Amount Paid"))
        layout.addWidget(self.amount_paid_input)
        layout.addWidget(QLabel("Change"))
        layout.addWidget(self.change_input)
        layout.addWidget(QLabel("Payment Type"))
        layout.addWidget(self.payment_type_input)

        add_button = QPushButton("Add")
        add_button.clicked.connect(self.add_invoice)
        layout.addWidget(add_button)

        self.setLayout(layout)

    def open_customers_window(self, event):
        self.customers_window = CustomersWindow(self)
        self.customers_window.exec()

    def open_firm_window(self, event):
        self.firm_window = FirmWindow(self)
        self.firm_window.exec()

    def open_products_window(self, event):
        self.products_window = ProductsWindow(self)
        self.products_window.exec()

    def update_change(self):
        try:
            total_amount = float(self.total_amount_input.text())
            amount_paid = float(self.amount_paid_input.text())
            change = amount_paid - total_amount
            self.change_input.setText(f"{change:.2f}")
        except ValueError:
            self.change_input.setText("0.00")

    def add_invoice(self):
        number = self.number_input.text()
        customer_ref_no = self.customer_ref_no_input.text()
        selling_company_ref_no = self.selling_company_ref_no_input.text()
        date = self.date_input.date().toString("yyyy-MM-dd")
        products = self.products_input.text()
        total_amount = self.total_amount_input.text()
        amount_paid = self.amount_paid_input.text()
        change = self.change_input.text()
        payment_type = self.payment_type_input.currentText()

        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        self.table.setItem(row_position, 0, QTableWidgetItem(number))
        self.table.setItem(row_position, 1, QTableWidgetItem(customer_ref_no))
        self.table.setItem(row_position, 2, QTableWidgetItem(selling_company_ref_no))
        self.table.setItem(row_position, 3, QTableWidgetItem(date))
        self.table.setItem(row_position, 4, QTableWidgetItem(products))
        self.table.setItem(row_position, 5, QTableWidgetItem(total_amount))
        self.table.setItem(row_position, 6, QTableWidgetItem(amount_paid))
        self.table.setItem(row_position, 7, QTableWidgetItem(change))
        self.table.setItem(row_position, 8, QTableWidgetItem(payment_type))

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
                    INSERT INTO invoices (number, customer_ref_no, selling_company_ref_no, date, products, 
                                          total_amount, amount_paid, change, payment_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
        number, customer_ref_no, selling_company_ref_no, date, products, total_amount, amount_paid, change,
        payment_type))
        connection.commit()
        connection.close()

        self.accept()

    def generate_invoice_number(self):
         connection = sqlite3.connect(DB_PATH)
         cursor = connection.cursor()
         cursor.execute('SELECT MAX(number) FROM invoices')
         result = cursor.fetchone()
         connection.close()
         if result[0] is not None:
             return str(int(result[0]) + 1).zfill(10)
         else:
             return "0000000001"

class EditInvoiceWindow(QDialog):
    def __init__(self, table, row):
        super().__init__()
        self.setWindowTitle("Edit Invoice")
        self.setGeometry(100, 100, 400, 500)
        self.table = table
        self.row = row

        layout = QVBoxLayout()

        self.number_input = QLineEdit()
        self.customer_ref_no_input = QLineEdit()
        self.selling_company_ref_no_input = QLineEdit()
        self.date_input = QLineEdit()
        self.products_input = QLineEdit()
        self.total_amount_input = QLineEdit()
        self.amount_paid_input = QLineEdit()
        self.change_input = QLineEdit()
        self.payment_type_input = QLineEdit()

        layout.addWidget(QLabel("Number"))
        layout.addWidget(self.number_input)
        layout.addWidget(QLabel("Customer Ref No"))
        layout.addWidget(self.customer_ref_no_input)
        layout.addWidget(QLabel("Selling Company Ref No"))
        layout.addWidget(self.selling_company_ref_no_input)
        layout.addWidget(QLabel("Date"))
        layout.addWidget(self.date_input)
        layout.addWidget(QLabel("Products"))
        layout.addWidget(self.products_input)
        layout.addWidget(QLabel("Total Amount"))
        layout.addWidget(self.total_amount_input)
        layout.addWidget(QLabel("Amount Paid"))
        layout.addWidget(self.amount_paid_input)
        layout.addWidget(QLabel("Change"))
        layout.addWidget(self.change_input)
        layout.addWidget(QLabel("Payment Type"))
        layout.addWidget(self.payment_type_input)

        self.load_invoice()

        edit_button = QPushButton("Edit")
        edit_button.clicked.connect(self.edit_invoice)
        layout.addWidget(edit_button)

        self.setLayout(layout)

    def load_invoice(self):
        self.number_input.setText(self.table.item(self.row, 0).text())
        self.customer_ref_no_input.setText(self.table.item(self.row, 1).text())
        self.selling_company_ref_no_input.setText(self.table.item(self.row, 2).text())
        self.date_input.setText(self.table.item(self.row, 3).text())
        self.products_input.setText(self.table.item(self.row, 4).text())
        self.total_amount_input.setText(self.table.item(self.row, 5).text())
        self.amount_paid_input.setText(self.table.item(self.row, 6).text())
        self.change_input.setText(self.table.item(self.row, 7).text())
        self.payment_type_input.setText(self.table.item(self.row, 8).text())

    def edit_invoice(self):
        number = self.number_input.text()
        customer_ref_no = self.customer_ref_no_input.text()
        selling_company_ref_no = self.selling_company_ref_no_input.text()
        date = self.date_input.text()
        products = self.products_input.text()
        total_amount = self.total_amount_input.text()
        amount_paid = self.amount_paid_input.text()
        change = self.change_input.text()
        payment_type = self.payment_type_input.text()

        self.table.setItem(self.row, 0, QTableWidgetItem(number))
        self.table.setItem(self.row, 1, QTableWidgetItem(customer_ref_no))
        self.table.setItem(self.row, 2, QTableWidgetItem(selling_company_ref_no))
        self.table.setItem(self.row, 3, QTableWidgetItem(date))
        self.table.setItem(self.row, 4, QTableWidgetItem(products))
        self.table.setItem(self.row, 5, QTableWidgetItem(total_amount))
        self.table.setItem(self.row, 6, QTableWidgetItem(amount_paid))
        self.table.setItem(self.row, 7, QTableWidgetItem(change))
        self.table.setItem(self.row, 8, QTableWidgetItem(payment_type))

        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('''
            UPDATE invoices
            SET customer_ref_no = ?, selling_company_ref_no = ?, date = ?, products = ?, 
                total_amount = ?, amount_paid = ?, change = ?, payment_type = ?
            WHERE number = ?
        ''', (customer_ref_no, selling_company_ref_no, date, products, total_amount, amount_paid, change, payment_type,
              number))
        connection.commit()
        connection.close()

        self.accept()

class CustomersWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Select Customer")
        self.setGeometry(100, 100, 400, 300)

        layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["ID", "Name"])
        layout.addWidget(self.table)

        self.load_customers()

        select_button = QPushButton("Select")
        select_button.clicked.connect(self.select_customer)
        layout.addWidget(select_button)

        self.setLayout(layout)

    def load_customers(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT id, name FROM customers')
        customers = cursor.fetchall()

        self.table.setRowCount(len(customers))
        for row_idx, customer in enumerate(customers):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(customer[0])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(customer[1]))

        connection.close()

    def select_customer(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            customer_id = self.table.item(current_row, 0).text()
            self.parent().customer_ref_no_input.setText(customer_id)
            self.accept()

# Update the ProductsWindow class:
# Update the ProductsWindow class:
class ProductsWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Select Products")
        self.setGeometry(100, 100, 600, 400)
        self.parent = parent

        layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["ID", "Name", "Quantity", "Unit Price"])
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        layout.addWidget(self.table)

        self.load_products()

        select_button = QPushButton("Select")
        select_button.clicked.connect(self.select_products)
        layout.addWidget(select_button)

        self.setLayout(layout)

    def load_products(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT id, name FROM products')
        products = cursor.fetchall()

        self.table.setRowCount(len(products))
        for row_idx, product in enumerate(products):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(product[0])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(product[1]))
            self.table.setItem(row_idx, 2, QTableWidgetItem("1"))  # Default quantity
            self.table.setItem(row_idx, 3, QTableWidgetItem("0.00"))  # Default unit price

        connection.close()

    def select_products(self):
        selected_products = []
        total_price = 0.0
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).isSelected():
                product_id = self.table.item(row, 0).text()
                product_name = self.table.item(row, 1).text()
                quantity = int(self.table.item(row, 2).text())
                unit_price = float(self.table.item(row, 3).text())
                total_price += quantity * unit_price
                selected_products.append(f"{product_name} (Qty: {quantity}, Unit Price: {unit_price:.2f})")

        self.parent.products_input.setText('; '.join(selected_products))
        self.parent.total_amount_input.setText(f"{total_price:.2f}")
        self.parent.amount_paid_input.setText(f"{total_price:.2f}")
        self.parent.change_input.setText("0.00")
        self.accept()


class FirmWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Select Firm")
        self.setGeometry(100, 100, 400, 300)

        layout = QVBoxLayout()

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["ID", "Name"])
        layout.addWidget(self.table)

        self.load_firms()

        select_button = QPushButton("Select")
        select_button.clicked.connect(self.select_firm)
        layout.addWidget(select_button)

        self.setLayout(layout)

    def load_firms(self):
        connection = sqlite3.connect(DB_PATH)
        cursor = connection.cursor()
        cursor.execute('SELECT id, name FROM firm')
        firms = cursor.fetchall()

        self.table.setRowCount(len(firms))
        for row_idx, firm in enumerate(firms):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(firm[0])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(firm[1]))

        connection.close()

    def select_firm(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            firm_id = self.table.item(current_row, 0).text()
            self.parent().selling_company_ref_no_input.setText(firm_id)
            self.accept()